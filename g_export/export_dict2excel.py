import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import globalvars

def dictexport(dicts):
    exporter = ExcelExporter(dicts)
    exporter.export_to_excel()

class ExcelExporter:
    def __init__(self, dicts):
        self.dicts = dicts
        # Get the current date and time
        self.new_folder = globalvars.output_path
        os.makedirs(self.new_folder, exist_ok=True)  # Create the new folder if it doesn't exist

    def export_to_excel(self):
        excel_filename = os.path.join(self.new_folder, "Dicts" + ".xlsx")
        # Try to open the existing workbook
        try:
            wb = openpyxl.load_workbook(excel_filename)
        except FileNotFoundError:
            # If the workbook does not exist, create a new one
            wb = Workbook()

        sheet = wb.active

        # Iterate over the dictionary items and write to Excel
        for row_index, (key, value) in enumerate(self.dicts.items(), start=1):
            # Write the key in column A
            sheet.cell(row=row_index, column=1, value=key)

            # Write the value(s) in column B
            if isinstance(value, (list, tuple)):
                # If the value is a list or tuple, convert it to a string
                sheet.cell(row=row_index, column=2, value=str(value))
            else:
                # If it's a single value, write it directly
                sheet.cell(row=row_index, column=2, value=value)

        # Save the Excel file
        wb.save(excel_filename)
