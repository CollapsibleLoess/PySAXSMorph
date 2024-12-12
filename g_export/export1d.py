import matplotlib.pyplot as plt
import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
import globalvars


def export_datas(dicts, x_label, y_label, log_scale=True, plot_type='scatter', clear_image=True):
    exporter = Exporter(dicts, x_label, y_label, log_scale, plot_type, clear_image)
    exporter.plot()
    if clear_image:
        plt.clf()  # 清除图像
    exporter.excel_exporter()


class Exporter:
    def __init__(self, dicts, x_label, y_label, log_scale, plot_type, clear_image):
        self.dicts = dicts
        self.x_label = x_label
        self.y_label = y_label
        self.log_scale = log_scale
        self.plot_type = plot_type
        self.save_plot = clear_image
        self.title = f"Plot {self.x_label} vs {self.y_label}"
        self.new_folder = globalvars.output_path
        os.makedirs(self.new_folder, exist_ok=True)  # Create the new folder if it doesn't exist

    def plot(self):
        # Extract x and y values from the list
        x_values = self.dicts[self.x_label]
        y_values = self.dicts[self.y_label]

        # Create the plot
        if self.log_scale:
            if self.plot_type == 'scatter':
                plt.loglog(x_values, y_values, 'o')
            elif self.plot_type == 'line':
                plt.loglog(x_values, y_values, 'r--')
        else:
            if self.plot_type == 'scatter':
                plt.scatter(x_values, y_values)
            elif self.plot_type == 'line':
                plt.plot(x_values, y_values, 'r--')

        # Set the labels
        plt.xlabel(self.x_label)
        plt.ylabel(self.y_label)
        plt.title(self.title)

        # Save the plot
        if self.save_plot:
            globalvars.plot_counter += 1
            plt.savefig(os.path.join(self.new_folder, str(globalvars.plot_counter) + self.title + ".png"))

    def excel_exporter(self):
        excel_filename = os.path.join(self.new_folder, "Lists" + ".xlsx")

        # Try to open the existing workbook
        try:
            wb = openpyxl.load_workbook(excel_filename)
        except FileNotFoundError:
            # If the workbook does not exist, create a new one
            wb = Workbook()

        sheet = wb.active

        # Find the first empty column
        empty_column = 1
        while sheet.cell(row=1, column=empty_column).value is not None:
            empty_column += 1

        if self.save_plot:
        # Write x_label to the first cell of the first column
            sheet.cell(row=1, column=empty_column, value=str(globalvars.plot_counter) + self.x_label)
        else:
            sheet.cell(row=1, column=empty_column, value=self.x_label)
        # Write x_values to the cells below x_label
        for i, x_value in enumerate(self.dicts[self.x_label]):
            sheet.cell(row=i + 2, column=empty_column, value=x_value)

        # Write y_label to the first cell of the first column
        sheet.cell(row=1, column=empty_column + 1, value=self.y_label)

        # Write y_values to the cells below x_label
        for i, y_value in enumerate(self.dicts[self.y_label]):  # This should be y_label, not x_label
            sheet.cell(row=i + 2, column=empty_column + 1, value=y_value)

        # Save the Excel file
        wb.save(excel_filename)
